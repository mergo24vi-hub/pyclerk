import pandas as pd
from openpyxl import load_workbook

# Файл з таблицею, яку оновлюємо
file_path = 'основна_таблиця.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# Читаємо джерело даних через pandas
df_source = pd.read_excel('джерело_вер2.xlsx')

# Зчитуємо заголовки (перший рядок)
headers = [cell.value for cell in ws[1]]
col_index = {name: idx + 1 for idx, name in enumerate(headers)}  # Excel columns = 1-based

# Вказуємо, який стовпець є унікальним для зв'язку
key_col = 'ПІБ'
# Список полів, які треба оновити
columns_to_update = ['конмоб','телефон', 'ддт', 'освіта']

# Проходимо по всіх рядках
for row in range(2, ws.max_row + 1):
    key_value = ws.cell(row=row, column=col_index[key_col]).value
    if pd.isna(key_value):
        continue

    match_row = df_source[df_source[key_col] == key_value]
    if not match_row.empty:
        for col_name in columns_to_update:
            col = col_index[col_name]
            current_cell = ws.cell(row=row, column=col)
            current_val = current_cell.value

            # Якщо клітинка порожня
            if current_val is None or str(current_val).strip() == '':
                new_val = match_row.iloc[0][col_name]
                if pd.notna(new_val):
                    current_cell.value = str(new_val)

# Зберігаємо у новий файл, оригінал не чіпаємо
wb.save('оновлена_таблиця.xlsx')
